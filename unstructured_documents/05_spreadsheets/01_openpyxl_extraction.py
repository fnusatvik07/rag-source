"""
Spreadsheet Extraction Method 1: openpyxl

openpyxl reads Excel files (.xlsx) with access to cells, sheets, formatting, and formulas.
It gives the most control over how spreadsheet data is extracted.

Best for: Cell-by-cell extraction, preserving sheet structure, handling merged cells and formatting.
"""

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

import openpyxl

from unstructured_documents.shared.chunking import (
    chunk_by_recursive_split,
    preview_chunks,
)

SAMPLE_DIR = Path(__file__).parent / "sample_docs"


def extract_all_sheets(xlsx_path: Path) -> dict[str, list[list]]:
    """Extract all sheets as lists of rows."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    sheets = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append(list(row))
        sheets[sheet_name] = rows
    return sheets


def extract_sheet_with_metadata(xlsx_path: Path) -> list[dict]:
    """Extract sheets with metadata (headers, dimensions, data types)."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    results = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        headers = [str(h) if h else f"col_{i}" for i, h in enumerate(rows[0])]
        data = rows[1:]

        results.append(
            {
                "sheet_name": sheet_name,
                "dimensions": ws.dimensions,
                "row_count": len(data),
                "headers": headers,
                "data": data,
            }
        )

    return results


def sheet_to_natural_language(sheet_name: str, headers: list, rows: list) -> str:
    """
    Convert a sheet to natural language description for RAG.
    This approach creates text that's more meaningful for embedding than raw tabular data.
    """
    lines = [f"## {sheet_name}\n"]
    lines.append(f"This sheet contains {len(rows)} records with the following fields: {', '.join(headers)}.\n")

    for row in rows:
        parts = []
        for header, value in zip(headers, row):
            if value is not None:
                parts.append(f"{header}: {value}")
        lines.append(f"- {'; '.join(parts)}")

    return "\n".join(lines)


def sheet_to_markdown_table(headers: list, rows: list) -> str:
    """Convert sheet data to a markdown table."""
    # Header row
    header_line = "| " + " | ".join(str(h) for h in headers) + " |"
    separator = "| " + " | ".join("---" for _ in headers) + " |"

    data_lines = []
    for row in rows:
        cells = [str(v) if v is not None else "" for v in row]
        data_lines.append("| " + " | ".join(cells) + " |")

    return "\n".join([header_line, separator] + data_lines)


if __name__ == "__main__":
    xlsx_path = SAMPLE_DIR / "multi_sheet.xlsx"

    # --- Basic extraction ---
    print("=" * 60)
    print("1. BASIC SHEET EXTRACTION")
    print("=" * 60)
    sheets = extract_all_sheets(xlsx_path)
    for name, rows in sheets.items():
        print(f"\nSheet: {name} ({len(rows)} rows)")
        for row in rows[:3]:
            print(f"  {row}")
        if len(rows) > 3:
            print(f"  ... and {len(rows) - 3} more rows")

    # --- Extraction with metadata ---
    print(f"\n{'=' * 60}")
    print("2. EXTRACTION WITH METADATA")
    print("=" * 60)
    sheet_meta = extract_sheet_with_metadata(xlsx_path)
    for sheet in sheet_meta:
        print(f"\nSheet: {sheet['sheet_name']}")
        print(f"  Dimensions: {sheet['dimensions']}")
        print(f"  Headers: {sheet['headers']}")
        print(f"  Row count: {sheet['row_count']}")

    # --- Natural language conversion (RAG-optimized) ---
    print(f"\n{'=' * 60}")
    print("3. NATURAL LANGUAGE CONVERSION (RAG-optimized)")
    print("=" * 60)
    for sheet in sheet_meta:
        nl_text = sheet_to_natural_language(sheet["sheet_name"], sheet["headers"], sheet["data"])
        print(f"\n{nl_text[:500]}")
        if len(nl_text) > 500:
            print("...")

    # --- Markdown table format ---
    print(f"\n{'=' * 60}")
    print("4. MARKDOWN TABLE FORMAT")
    print("=" * 60)
    for sheet in sheet_meta[:1]:  # Just show first sheet
        md_table = sheet_to_markdown_table(sheet["headers"], sheet["data"])
        print(f"\n{md_table}")

    # --- Chunking natural language output ---
    print(f"\n{'=' * 60}")
    print("5. CHUNKED OUTPUT FOR RAG")
    print("=" * 60)
    all_text = "\n\n".join(sheet_to_natural_language(s["sheet_name"], s["headers"], s["data"]) for s in sheet_meta)
    chunks = chunk_by_recursive_split(all_text, chunk_size=400)
    preview_chunks(chunks)
