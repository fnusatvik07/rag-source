"""Generate sample spreadsheet files for testing extraction methods."""

import csv
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill

SAMPLE_DIR = Path(__file__).parent


def generate_multi_sheet_workbook():
    """Generate an XLSX file with multiple sheets and formatting."""
    wb = openpyxl.Workbook()

    # Sheet 1: Employee Directory
    ws1 = wb.active
    ws1.title = "Employee Directory"
    headers = ["ID", "Name", "Department", "Title", "Email", "Salary", "Start Date"]
    ws1.append(headers)
    for cell in ws1[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    employees = [
        [
            101,
            "Alice Johnson",
            "Engineering",
            "Senior Developer",
            "alice@company.com",
            125000,
            "2020-03-15",
        ],
        [
            102,
            "Bob Smith",
            "Marketing",
            "Marketing Manager",
            "bob@company.com",
            95000,
            "2019-07-01",
        ],
        [
            103,
            "Carol Williams",
            "Engineering",
            "Tech Lead",
            "carol@company.com",
            145000,
            "2018-01-10",
        ],
        [
            104,
            "David Brown",
            "Sales",
            "Account Executive",
            "david@company.com",
            85000,
            "2021-06-20",
        ],
        [
            105,
            "Eve Davis",
            "Engineering",
            "Junior Developer",
            "eve@company.com",
            75000,
            "2023-02-01",
        ],
        [
            106,
            "Frank Miller",
            "HR",
            "HR Director",
            "frank@company.com",
            110000,
            "2017-11-15",
        ],
        [
            107,
            "Grace Lee",
            "Finance",
            "Financial Analyst",
            "grace@company.com",
            90000,
            "2022-04-01",
        ],
        [
            108,
            "Henry Wilson",
            "Engineering",
            "DevOps Engineer",
            "henry@company.com",
            120000,
            "2020-09-10",
        ],
    ]
    for emp in employees:
        ws1.append(emp)

    # Sheet 2: Quarterly Revenue
    ws2 = wb.create_sheet("Quarterly Revenue")
    ws2.append(["Quarter", "Product A", "Product B", "Product C", "Total"])
    for cell in ws2[1]:
        cell.font = Font(bold=True)

    revenue = [
        ["Q1 2024", 150000, 95000, 45000, 290000],
        ["Q2 2024", 175000, 110000, 52000, 337000],
        ["Q3 2024", 190000, 125000, 68000, 383000],
        ["Q4 2024", 220000, 140000, 75000, 435000],
    ]
    for row in revenue:
        ws2.append(row)

    # Add a summary row
    ws2.append(["Total", 735000, 470000, 240000, 1445000])
    for cell in ws2[6]:
        cell.font = Font(bold=True)

    # Sheet 3: Project Tracker
    ws3 = wb.create_sheet("Project Tracker")
    ws3.append(["Project", "Status", "Owner", "Priority", "Deadline", "Budget", "Notes"])
    for cell in ws3[1]:
        cell.font = Font(bold=True)

    projects = [
        [
            "Website Redesign",
            "In Progress",
            "Alice",
            "High",
            "2025-03-01",
            50000,
            "Phase 2 of 3 complete",
        ],
        [
            "Mobile App v2",
            "Planning",
            "Carol",
            "High",
            "2025-06-15",
            120000,
            "Requirements gathering",
        ],
        [
            "Data Pipeline",
            "Complete",
            "Henry",
            "Medium",
            "2025-01-15",
            35000,
            "Deployed to production",
        ],
        [
            "CRM Integration",
            "In Progress",
            "Bob",
            "Medium",
            "2025-04-01",
            25000,
            "API development phase",
        ],
        [
            "Security Audit",
            "Not Started",
            "Frank",
            "High",
            "2025-02-28",
            15000,
            "Vendor selection pending",
        ],
    ]
    for proj in projects:
        ws3.append(proj)

    wb.save(SAMPLE_DIR / "multi_sheet.xlsx")
    print("Generated: multi_sheet.xlsx")


def generate_csv_files():
    """Generate sample CSV files with different formats."""
    # Simple CSV
    with open(SAMPLE_DIR / "simple_data.csv", "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["City", "Country", "Population", "Area_km2", "Density"])
        data = [
            ["Tokyo", "Japan", 13960000, 2194, 6363],
            ["Delhi", "India", 11030000, 1484, 7433],
            ["Shanghai", "China", 24870000, 6341, 3924],
            ["São Paulo", "Brazil", 12330000, 1521, 8105],
            ["Mexico City", "Mexico", 9210000, 1485, 6202],
            ["Cairo", "Egypt", 9540000, 3085, 3093],
            ["Mumbai", "India", 12440000, 603, 20634],
            ["London", "UK", 8982000, 1572, 5713],
        ]
        writer.writerows(data)
    print("Generated: simple_data.csv")

    # CSV with mixed content (text descriptions)
    with open(SAMPLE_DIR / "products.csv", "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["SKU", "Product Name", "Category", "Description", "Price", "Stock"])
        products = [
            [
                "WDG-001",
                "Smart Widget Pro",
                "Electronics",
                (
                    "Advanced smart widget with WiFi connectivity, 4K display, and voice control."
                    " Compatible with all major smart home ecosystems."
                ),
                299.99,
                150,
            ],
            [
                "TBL-002",
                "Ergonomic Desk",
                "Furniture",
                (
                    "Height-adjustable standing desk with memory presets."
                    " Supports up to 200 lbs. Available in walnut and oak finishes."
                ),
                549.99,
                42,
            ],
            [
                "SFT-003",
                "CloudSync Suite",
                "Software",
                (
                    "Enterprise file synchronization and collaboration platform."
                    " Includes 1TB storage, version control, and real-time editing."
                ),
                19.99,
                999,
            ],
            [
                "ACC-004",
                "USB-C Hub Deluxe",
                "Accessories",
                "12-in-1 USB-C hub with dual HDMI, ethernet, SD card reader, and 100W passthrough charging.",
                79.99,
                320,
            ],
        ]
        writer.writerows(products)
    print("Generated: products.csv")


if __name__ == "__main__":
    generate_multi_sheet_workbook()
    generate_csv_files()
    print("\nAll spreadsheet samples generated!")
